"""Export test cases to Excel (XLSX) with professional formatting."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from testgen.config import GenerationResult


# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2C5282")
BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

PRIORITY_FILLS = {
    "High": PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
    "Medium": PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
    "Low": PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
}

ALT_ROW_FILL = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")


def export_xlsx(result: GenerationResult, output_path: str | Path) -> Path:
    """Export generation result to a professional XLSX file.

    Creates three sheets:
    - Summary: overview statistics
    - Test Cases: all test cases with details and auto-filters
    - Traceability Matrix: requirement to test case mapping

    Args:
        result: The generation result containing test cases.
        output_path: Path for the output XLSX file.

    Returns:
        Path to the created file.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _create_summary_sheet(wb, result)
    _create_test_cases_sheet(wb, result)
    _create_traceability_sheet(wb, result)

    wb.save(str(path))
    return path


def _create_summary_sheet(wb: Workbook, result: GenerationResult) -> None:
    """Create the Summary sheet with overview statistics."""
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "Test Cases Report — Summary"
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    summary = result.summary
    row = 3

    # Overview
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "Overview"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    overview_data = [
        ("Total Test Cases", summary["total"]),
        ("Requirements Covered", summary["requirements_count"]),
    ]
    for label, value in overview_data:
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = BODY_FONT
        ws[f"B{row}"].value = value
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=12, color="2C5282")
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # By Priority
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "By Priority"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    for col_label, col_letter in [("Priority", "A"), ("Count", "B")]:
        cell = ws[f"{col_letter}{row}"]
        cell.value = col_label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1

    for priority, count in summary.get("by_priority", {}).items():
        ws[f"A{row}"].value = priority
        ws[f"A{row}"].font = BODY_FONT
        ws[f"A{row}"].border = THIN_BORDER
        if priority in PRIORITY_FILLS:
            ws[f"A{row}"].fill = PRIORITY_FILLS[priority]
        ws[f"B{row}"].value = count
        ws[f"B{row}"].font = BODY_FONT
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = THIN_BORDER
        row += 1

    row += 1

    # By Type
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "By Type"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    for col_label, col_letter in [("Type", "A"), ("Count", "B")]:
        cell = ws[f"{col_letter}{row}"]
        cell.value = col_label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1

    for test_type, count in summary.get("by_type", {}).items():
        ws[f"A{row}"].value = test_type
        ws[f"A{row}"].font = BODY_FONT
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"].value = count
        ws[f"B{row}"].font = BODY_FONT
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = THIN_BORDER
        row += 1

    row += 1

    # By Technique
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "By ISTQB Technique"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    for col_label, col_letter in [("Technique", "A"), ("Count", "B")]:
        cell = ws[f"{col_letter}{row}"]
        cell.value = col_label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1

    for technique, count in summary.get("by_technique", {}).items():
        ws[f"A{row}"].value = technique
        ws[f"A{row}"].font = BODY_FONT
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"].value = count
        ws[f"B{row}"].font = BODY_FONT
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


def _create_test_cases_sheet(wb: Workbook, result: GenerationResult) -> None:
    """Create the Test Cases sheet with all details and auto-filters."""
    ws = wb.create_sheet("Test Cases")

    is_ser = result.step_format == "step_expected"
    if is_ser:
        headers = [
            "ID", "Title", "Preconditions", "Step", "Expected",
            "Test Data", "Expected Result", "Priority",
            "Type", "ISTQB Technique", "Traceability",
        ]
        column_widths = [8, 35, 25, 35, 35, 25, 30, 10, 12, 22, 12]
    else:
        headers = [
            "ID", "Title", "Preconditions", "Steps (Given)", "Steps (When)",
            "Steps (Then)", "Test Data", "Expected Result", "Priority",
            "Type", "ISTQB Technique", "Traceability",
        ]
        column_widths = [8, 35, 25, 30, 30, 30, 25, 30, 10, 12, 22, 12]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 25

    # Set column widths
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, tc in enumerate(result.test_cases, 2):
        # Combine steps into readable strings
        if is_ser:
            step_parts = []
            expected_parts = []
            for i, step in enumerate(tc.steps, 1):
                prefix = f"{i}. " if len(tc.steps) > 1 else ""
                step_parts.append(f"{prefix}{step.step}")
                expected_parts.append(f"{prefix}{step.expected}")
            row_data = [
                tc.id, tc.title, tc.preconditions,
                "\n".join(step_parts), "\n".join(expected_parts),
                tc.test_data, tc.expected_result,
                tc.priority.value, tc.type.value,
                tc.istqb_technique.value, tc.traceability,
            ]
        else:
            given_parts = []
            when_parts = []
            then_parts = []
            for i, step in enumerate(tc.steps, 1):
                prefix = f"{i}. " if len(tc.steps) > 1 else ""
                given_parts.append(f"{prefix}{step.given}")
                when_parts.append(f"{prefix}{step.when}")
                then_parts.append(f"{prefix}{step.then}")
            row_data = [
                tc.id, tc.title, tc.preconditions,
                "\n".join(given_parts), "\n".join(when_parts), "\n".join(then_parts),
                tc.test_data, tc.expected_result,
                tc.priority.value, tc.type.value,
                tc.istqb_technique.value, tc.traceability,
            ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER

            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

            # Priority coloring
            priority_col = 8 if is_ser else 9
            if col_idx == priority_col and value in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[value]
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row_idx].height = max(30, 15 * max(len(tc.steps), 1))

    # Auto-filter
    if result.test_cases:
        last_col = get_column_letter(len(headers))
        last_row = len(result.test_cases) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Freeze top row
    ws.freeze_panes = "A2"


def _create_traceability_sheet(wb: Workbook, result: GenerationResult) -> None:
    """Create the Traceability Matrix sheet."""
    ws = wb.create_sheet("Traceability Matrix")

    headers = ["Requirement ID", "Requirement Text", "Test Cases", "Count"]
    column_widths = [15, 50, 40, 10]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 25

    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Build traceability map
    trace_map: dict[str, list[str]] = {}
    for tc in result.test_cases:
        if tc.traceability not in trace_map:
            trace_map[tc.traceability] = []
        trace_map[tc.traceability].append(tc.id)

    # Write data
    row_idx = 2
    for req in result.requirements:
        tc_ids = trace_map.get(req.id, [])
        row_data = [
            req.id,
            req.text,
            ", ".join(tc_ids),
            len(tc_ids),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER

            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        row_idx += 1

    # Auto-filter
    if result.requirements:
        last_row = len(result.requirements) + 1
        ws.auto_filter.ref = f"A1:D{last_row}"

    ws.freeze_panes = "A2"
