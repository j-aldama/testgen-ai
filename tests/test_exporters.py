"""Tests for all 4 exporters."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from testgen.config import GenerationResult
from testgen.exporters.json_exporter import export_json, to_json_string
from testgen.exporters.markdown_exporter import export_markdown, render_markdown
from testgen.exporters.xlsx_exporter import export_xlsx


class TestJsonExporter:
    """Tests for JSON exporter."""

    def test_export_json_creates_file(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        output = tmp_path / "test_cases.json"
        result_path = export_json(sample_generation_result, output)

        assert result_path.exists()
        assert result_path.suffix == ".json"

    def test_export_json_valid_structure(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        output = tmp_path / "test_cases.json"
        export_json(sample_generation_result, output)

        data = json.loads(output.read_text(encoding="utf-8"))
        assert "session_id" in data
        assert "summary" in data
        assert "requirements" in data
        assert "test_cases" in data
        assert len(data["test_cases"]) == 3

    def test_export_json_test_case_fields(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        output = tmp_path / "test_cases.json"
        export_json(sample_generation_result, output)

        data = json.loads(output.read_text(encoding="utf-8"))
        tc = data["test_cases"][0]
        required_fields = [
            "id", "title", "preconditions", "steps", "test_data",
            "expected_result", "priority", "type", "istqb_technique", "traceability",
        ]
        for field in required_fields:
            assert field in tc, f"Missing field: {field}"

    def test_to_json_string(self, sample_generation_result: GenerationResult) -> None:
        result = to_json_string(sample_generation_result)
        data = json.loads(result)
        assert data["session_id"] == "test-session-123"
        assert len(data["test_cases"]) == 3

    def test_export_json_creates_parent_dirs(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        output = tmp_path / "deep" / "nested" / "test_cases.json"
        result_path = export_json(sample_generation_result, output)
        assert result_path.exists()


class TestMarkdownExporter:
    """Tests for Markdown exporter."""

    def test_export_markdown_creates_file(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        output = tmp_path / "test_cases.md"
        result_path = export_markdown(sample_generation_result, output)

        assert result_path.exists()
        assert result_path.suffix == ".md"

    def test_markdown_contains_title(
        self, sample_generation_result: GenerationResult
    ) -> None:
        content = render_markdown(sample_generation_result)
        assert "# Test Cases Report" in content

    def test_markdown_contains_summary(
        self, sample_generation_result: GenerationResult
    ) -> None:
        content = render_markdown(sample_generation_result)
        assert "## Summary" in content
        assert "Total Test Cases" in content
        assert "3" in content  # 3 test cases

    def test_markdown_contains_test_case_details(
        self, sample_generation_result: GenerationResult
    ) -> None:
        content = render_markdown(sample_generation_result)
        assert "TC-001" in content
        assert "TC-002" in content
        assert "TC-003" in content
        assert "Given:" in content
        assert "When:" in content
        assert "Then:" in content

    def test_markdown_contains_traceability_matrix(
        self, sample_generation_result: GenerationResult
    ) -> None:
        content = render_markdown(sample_generation_result)
        assert "Traceability Matrix" in content
        assert "REQ-001" in content

    def test_markdown_contains_overview_table(
        self, sample_generation_result: GenerationResult
    ) -> None:
        content = render_markdown(sample_generation_result)
        assert "Test Cases Overview" in content
        assert "| ID | Title |" in content


class TestXlsxExporter:
    """Tests for Excel exporter."""

    def test_export_xlsx_creates_file(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        output = tmp_path / "test_cases.xlsx"
        result_path = export_xlsx(sample_generation_result, output)

        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

    def test_xlsx_has_three_sheets(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Test Cases" in sheet_names
        assert "Traceability Matrix" in sheet_names

    def test_xlsx_summary_sheet_has_data(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        ws = wb["Summary"]
        # Title should be in A1
        assert ws["A1"].value is not None
        assert "Summary" in str(ws["A1"].value)

    def test_xlsx_test_cases_sheet_has_headers(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        ws = wb["Test Cases"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 13)]
        assert "ID" in headers
        assert "Title" in headers
        assert "Priority" in headers
        assert "Type" in headers
        assert "ISTQB Technique" in headers

    def test_xlsx_test_cases_has_correct_row_count(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        ws = wb["Test Cases"]
        # Header + 3 test cases = 4 rows with data
        assert ws.cell(row=2, column=1).value == "TC-001"
        assert ws.cell(row=3, column=1).value == "TC-002"
        assert ws.cell(row=4, column=1).value == "TC-003"

    def test_xlsx_has_auto_filter(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        ws = wb["Test Cases"]
        assert ws.auto_filter.ref is not None

    def test_xlsx_traceability_matrix(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        ws = wb["Traceability Matrix"]
        # Should have requirements as rows
        assert ws.cell(row=2, column=1).value == "REQ-001"

    def test_xlsx_header_styling(
        self, sample_generation_result: GenerationResult, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        output = tmp_path / "test_cases.xlsx"
        export_xlsx(sample_generation_result, output)

        wb = load_workbook(str(output))
        ws = wb["Test Cases"]
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color is not None
